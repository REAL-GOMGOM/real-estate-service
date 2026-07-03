"""
확장 144개 지역 종합 분석 Excel + 학군 실측 교정
- Sheet1: 전체 144개 지역 (신규 20개 편입 강조)
- Sheet2: 신규 편입 지역 상세 (매매지수·교통·평당가)
- Sheet3: 학군점수 실측 교정 대조
- Sheet4: 광역권 요약 (신규 광역권 포함)
- Sheet5: 방법론
"""
import sys
sys.path.insert(0, '/home/claude')
from scoring_model import build_dataframe, compute_scores, sensitivity_analysis, WEIGHT_SCENARIOS
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

df = build_dataframe()
scored = compute_scores(df, weights=WEIGHT_SCENARIOS['Base'])

wb = Workbook()
thin = Border(*[Side(style='thin', color='D0D0D0')]*4)

def hstyle(c, color='1F4E78'):
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    c.fill = PatternFill('solid', start_color=color)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin

# 신규 편입 지역 번호 (125~144)
NEW_REGIONS = set(range(125, 145))

# ============================================================
# Sheet 1: 전체 144개
# ============================================================
ws1 = wb.active
ws1.title = "전체144_지역"

headers = ["No", "지역명", "그룹", "원본점수", "평당가\n(만원)", "2025누계\n매매%", 
           "2026최신\n주간%", "교통", "학군", "산업", "공급", "전세가율", 
           "재산정점수", "점수차", "구분"]
for i, h in enumerate(headers, 1):
    hstyle(ws1.cell(row=1, column=i, value=h))

group_colors = {
    '서울': 'FFF2CC', '인천': 'FFE4E1', '성남': 'E2EFDA', '수원': 'E2EFDA',
    '고양': 'E2EFDA', '용인': 'E2EFDA', '안양': 'E2EFDA', '안산': 'E2EFDA',
    '과천': 'E2EFDA', '광명': 'E2EFDA', '하남': 'E2EFDA', '의왕': 'E2EFDA',
    '부천': 'E2EFDA', '구리': 'E2EFDA', '남양주': 'E2EFDA', '의정부': 'E2EFDA',
    '시흥': 'E2EFDA', '김포': 'E2EFDA', '화성': 'E2EFDA', '평택': 'E2EFDA',
    '파주': 'E2EFDA', '군포': 'E2EFDA', '신도시': 'DDEBF7', '2기': 'DDEBF7',
    '3기': 'FCE4D6', '기타': 'FCE4D6', '부산': 'F4CCCC', '대구': 'D9D9D9', 
    '울산': 'E6B8AF',
    # 신규 광역권
    '대전': 'D5E8D4', '광주': 'D4E1F5', '세종': 'FFE6CC', '전북': 'F8CECC',
    '경남': 'E1D5E7', '충북': 'FFF2CC', '충남': 'D5E8D4',
}

for _, r in scored.iterrows():
    no = int(r['no'])
    is_new = no in NEW_REGIONS
    vals = [
        no, r['name'], r['region'], r['original_score'],
        r['평당가'] if pd.notna(r['평당가']) else '분양전',
        r['매매누계_2025'], r['주간지수_2026'],
        r['교통'], r['학군'], r['산업'], r['공급'],
        r['전세가율'] if pd.notna(r['전세가율']) else '-',
        r['재산정점수'], r['점수차'],
        '🆕신규' if is_new else '기존'
    ]
    row_i = no + 1
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=row_i, column=col, value=v)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col == 3:
            c.fill = PatternFill('solid', start_color=group_colors.get(r['region'], 'FFFFFF'))
        if col == 13:  # 재산정점수
            s = r['재산정점수']
            if s < 2.0: c.fill = PatternFill('solid', start_color='9ACD32')
            elif s < 2.5: c.fill = PatternFill('solid', start_color='B5E27C')
            elif s < 3.0: c.fill = PatternFill('solid', start_color='D4F0A0')
            elif s < 3.5: c.fill = PatternFill('solid', start_color='FFF2CC')
            elif s < 4.0: c.fill = PatternFill('solid', start_color='FFCC99')
            else: c.fill = PatternFill('solid', start_color='FF9999')
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 15 and is_new:  # 신규 강조
            c.fill = PatternFill('solid', start_color='FFEB9C')
            c.font = Font(name='Arial', size=9, bold=True, color='C55A11')

widths1 = {'A': 4, 'B': 13, 'C': 8, 'D': 8, 'E': 10, 'F': 9, 'G': 9,
           'H': 6, 'I': 6, 'J': 6, 'K': 6, 'L': 9, 'M': 10, 'N': 8, 'O': 8}
for col, w in widths1.items():
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 36
ws1.freeze_panes = 'C2'
ws1.auto_filter.ref = f"A1:O{len(scored)+1}"

# ============================================================
# Sheet 2: 신규 편입 지역 상세
# ============================================================
ws2 = wb.create_sheet("신규편입_20개_상세")

new_detail = [
    # 지역, 광역권, 평당가, 2025누계, 2026최신주간, 교통현황, 학군특징, 재산정점수
    ("대전 서구", "대전", "1,611", "-1.0%", "+0.04%", "1호선·대전역세권개발", "둔산동 대전 최고학군(학원438)", None),
    ("대전 유성구", "대전", "1,807", "+0.5%", "+0.02%", "대전1호선·트램2호선 공사", "전민동·대덕연구단지 배후", None),
    ("대전 중구", "대전", "1,300", "-2.0%", "-0.01%", "1호선", "원도심", None),
    ("대전 동구", "대전", "1,150", "-2.5%", "-0.01%", "대전역세권 개발", "원도심 재개발", None),
    ("대전 대덕구", "대전", "1,100", "-2.5%", "-0.01%", "신탄진·경부선", "산업단지 배후", None),
    ("광주 남구", "광주", "1,700", "-0.5%", "+0.03%", "지하철1호선", "봉선동 광주 최고학군·부촌", None),
    ("광주 광산구", "광주", "1,450", "-0.3%", "+0.04%", "지하철1호선·수완지구", "수완·첨단 신도시", None),
    ("광주 서구", "광주", "1,550", "-0.5%", "+0.02%", "상무지구", "상무지구 행정타운", None),
    ("광주 북구", "광주", "1,350", "-0.8%", "+0.02%", "지하철1호선", "운암·용봉동", None),
    ("광주 동구", "광주", "1,400", "-1.0%", "+0.01%", "지하철1호선", "원도심·재개발", None),
    ("세종시", "세종", "2,200", "+2.9%", "0.00%", "BRT·GTX 미정·정부청사", "종촌·아름동, 2025 +2.92%급등", None),
    ("전주 완산구", "전북", "1,146", "+1.0%", "+0.03%", "전주역·호남선", "효자동 서울대진학8%(최상위)", None),
    ("전주 덕진구", "전북", "1,200", "+1.5%", "+0.03%", "전주역·에코시티", "송천동 에코시티 신도시", None),
    ("창원 성산구", "경남", "1,500", "+8.0%", "+0.15%", "창원터널·마산선", "상남동 창원 대장, 기계산업", None),
    ("창원 의창구", "경남", "1,452", "+9.0%", "+0.14%", "창원중앙역", "용호동, 2025 +11.2%급등", None),
    ("김해시", "경남", "1,300", "+2.0%", "+0.08%", "부산김해경전철·율하", "율하·장유신도시, 부산배후", None),
    ("아산시", "충남", "1,400", "+1.5%", "+0.05%", "KTX천안아산역·1호선", "탕정 특목고42.6%, 삼성벨트", None),
    ("청주 흥덕구", "충북", "1,600", "+1.5%", "+0.06%", "오송역KTX·오송바이오", "복대동 지웰시티3,350, 반도체", None),
    ("청주 상당구", "충북", "1,400", "+1.0%", "+0.05%", "청주역·방서지구", "방서지구 신도시", None),
    ("송도국제도시", "인천", "3,700", "-3.0%", "+0.03%", "인천1호선·GTX-B예정", "송도동 신흥학군, 대장단지폭락", None),
]

# 재산정점수 매핑
score_map = dict(zip(scored['name'], scored['재산정점수']))
name_key = {
    "대전 서구": "대전서구", "대전 유성구": "대전유성구", "대전 중구": "대전중구",
    "대전 동구": "대전동구", "대전 대덕구": "대전대덕구",
    "광주 남구": "광주남구", "광주 광산구": "광주광산구", "광주 서구": "광주서구",
    "광주 북구": "광주북구", "광주 동구": "광주동구",
    "세종시": "세종시", "전주 완산구": "전주완산구", "전주 덕진구": "전주덕진구",
    "창원 성산구": "창원성산구", "창원 의창구": "창원의창구", "김해시": "김해시",
    "아산시": "아산시", "청주 흥덕구": "청주흥덕구", "청주 상당구": "청주상당구",
    "송도국제도시": "송도국제도시",
}

nh = ["지역", "광역권", "평당가\n(만원)", "2025\n누계%", "2026최신\n주간%", 
      "교통 현황", "학군 특징", "재산정\n점수"]
for i, h in enumerate(nh, 1):
    hstyle(ws2.cell(row=1, column=i, value=h), color='C55A11')

for idx, row in enumerate(new_detail, 2):
    region_name = name_key.get(row[0], row[0])
    rescore = score_map.get(region_name, '-')
    vals = list(row[:7]) + [round(rescore, 2) if isinstance(rescore, (int, float)) else '-']
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=idx, column=col, value=v)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col in [3, 4, 5, 8]:
            c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        if col == 1:
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 8 and isinstance(v, (int, float)):
            if v < 3.0: c.fill = PatternFill('solid', start_color='D4F0A0')
            elif v < 3.5: c.fill = PatternFill('solid', start_color='FFF2CC')
            elif v < 4.0: c.fill = PatternFill('solid', start_color='FFCC99')
            else: c.fill = PatternFill('solid', start_color='FF9999')

widths2 = {'A': 15, 'B': 8, 'C': 10, 'D': 8, 'E': 9, 'F': 24, 'G': 30, 'H': 8}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 32
for r in range(2, len(new_detail)+2):
    ws2.row_dimensions[r].height = 40

# ============================================================
# Sheet 3: 학군점수 실측 교정 대조
# ============================================================
ws3 = wb.create_sheet("학군점수_실측교정")

school_recheck = [
    # 지역, 기존모델학군(0~10), 실측학군(0~10), 실측서울대%, 실측특목%, 학원수, 판정
    ("대치동(강남)", 10.0, 6.52, "3.75", "27.13", "1442", "실측 최상위 유지"),
    ("반포동(서초)", 9.8, 6.57, "4.20", "36.02", "456", "실측 전국1위"),
    ("목동(양천)", 9.5, 5.59, "2.41", "24.65", "1022", "약간 과대"),
    ("중계동(노원)", 8.5, 3.84, "1.07", "9.75", "597", "과대(학원많으나 진학낮음)"),
    ("분당 정자동", 9.5, 4.95, "2.26", "21.96", "455", "과대"),
    ("평촌(동안구)", 9.5, 3.63, "0.66", "13.65", "425", "대폭 과대"),
    ("수성구 범어동", 9.5, 3.65, "0.99", "6.84", "616", "대폭 과대"),
    ("해운대 우동", 9.0, 3.83, "0.99", "18.75", "306", "과대"),
    ("일산 마두동", 9.0, 3.27, "0.45", "11.86", "412", "대폭 과대"),
    # 신규 지역 학군 실측
    ("대전 둔산동", 9.0, 4.30, "1.62", "17.46", "438", "🆕지방 최고학군급"),
    ("부산 구서동", "-", 4.66, "1.75", "29.66", "266", "🆕부산 최고학군"),
    ("광주 봉선동", "-", 2.72, "0.43", "3.90", "358", "🆕광주 최고학군"),
    ("전주 효자동", "-", 5.67, "8.02", "3.28", "195", "🆕서울대진학 전국최고"),
    ("아산 탕정면", "-", 4.31, "2.65", "42.59", "29", "🆕특목고진학 최상위"),
    ("송도동(연수)", 8.0, 3.68, "0.68", "17.47", "467", "🆕신흥학군"),
]

rh = ["지역", "기존모델\n학군점수", "실측\n학군점수", "서울대\n진학%", "특목고\n진학%", 
      "학원수", "판정"]
for i, h in enumerate(rh, 1):
    hstyle(ws3.cell(row=1, column=i, value=h), color='548235')

for idx, row in enumerate(school_recheck, 2):
    for col, v in enumerate(row, 1):
        c = ws3.cell(row=idx, column=col, value=v)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col == 1:
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 3:  # 실측 학군점수
            if isinstance(v, (int, float)):
                if v >= 5: c.fill = PatternFill('solid', start_color='B5E27C')
                elif v >= 4: c.fill = PatternFill('solid', start_color='D4F0A0')
                else: c.fill = PatternFill('solid', start_color='FFF2CC')
        if col == 7:
            if '대폭' in str(v):
                c.fill = PatternFill('solid', start_color='FFCCCC')
                c.font = Font(name='Arial', size=9, color='C00000', bold=True)
            elif '🆕' in str(v):
                c.fill = PatternFill('solid', start_color='FFEB9C')
                c.font = Font(name='Arial', size=9, color='C55A11', bold=True)

widths3 = {'A': 14, 'B': 9, 'C': 9, 'D': 8, 'E': 8, 'F': 7, 'G': 24}
for col, w in widths3.items():
    ws3.column_dimensions[col].width = w
ws3.row_dimensions[1].height = 32
for r in range(2, len(school_recheck)+2):
    ws3.row_dimensions[r].height = 26

# ============================================================
# Sheet 4: 광역권 요약 (신규 포함)
# ============================================================
ws4 = wb.create_sheet("광역권_요약통계")

# 광역권 대분류 매핑
def macro_region(region):
    seoul_gyeonggi = ['서울', '인천', '성남', '수원', '고양', '용인', '안양', '안산',
                      '과천', '광명', '하남', '의왕', '부천', '구리', '남양주', '의정부',
                      '시흥', '김포', '화성', '평택', '파주', '군포', '신도시', '2기', 
                      '3기', '기타']
    if region in seoul_gyeonggi: return '수도권'
    if region == '부산': return '부산'
    if region == '대구': return '대구'
    if region == '울산': return '울산'
    if region == '대전': return '대전(신규)'
    if region == '광주': return '광주(신규)'
    if region == '세종': return '세종(신규)'
    if region in ['전북']: return '전북(신규)'
    if region in ['경남']: return '경남(신규)'
    if region in ['충북']: return '충북(신규)'
    if region in ['충남']: return '충남(신규)'
    return '기타'

scored['광역권'] = scored['region'].apply(macro_region)
valid = scored[scored['평당가'].notna()].copy()

summary = valid.groupby('광역권').agg({
    '평당가': ['mean', 'min', 'max'],
    '매매누계_2025': 'mean',
    '주간지수_2026': 'mean',
    '재산정점수': 'mean',
    'no': 'count'
}).round(2)
summary.columns = ['평당가평균', '평당가최저', '평당가최고', '2025누계평균',
                   '2026주간평균', '재산정점수평균', '지역수']
summary = summary.sort_values('평당가평균', ascending=False)

headers4 = ['광역권'] + list(summary.columns)
for i, h in enumerate(headers4, 1):
    hstyle(ws4.cell(row=1, column=i, value=h), color='4472C4')

for r_idx, (region, row) in enumerate(summary.iterrows(), 2):
    vals = [region] + list(row.values)
    for col, v in enumerate(vals, 1):
        c = ws4.cell(row=r_idx, column=col, value=v)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col == 1:
            c.font = Font(name='Arial', size=9, bold=True)
            if '신규' in str(region):
                c.fill = PatternFill('solid', start_color='FFEB9C')

for col_letter in 'ABCDEFGH':
    ws4.column_dimensions[col_letter].width = 13
ws4.column_dimensions['A'].width = 14
ws4.row_dimensions[1].height = 30

# ============================================================
# Sheet 5: 방법론
# ============================================================
ws5 = wb.create_sheet("방법론_확장")
method = [
    ["항목", "내용"],
    ["기준일", "2026년 7월 3일"],
    ["총 지역수", "144개 (기존 124 + 신규 편입 20)"],
    ["", ""],
    ["===== 신규 편입 20개 지역 =====", ""],
    ["대전 (5개구)", "서구·유성구·중구·동구·대덕구"],
    ["광주 (5개구)", "남구·광산구·서구·북구·동구"],
    ["세종 (1)", "세종특별자치시"],
    ["전북 전주 (2개구)", "완산구·덕진구"],
    ["경남 (3)", "창원 성산구·의창구, 김해시"],
    ["충남 (1)", "아산시"],
    ["충북 청주 (2개구)", "흥덕구·상당구"],
    ["인천 (1)", "송도국제도시(연수구 송도동 특화)"],
    ["", ""],
    ["===== 편입 사유 =====", ""],
    ["발견 경위", "학군 조사 중 기존 리스트에 광역권 자체가 누락된 것 발견"],
    ["대전·광주·세종", "5대 광역시·특별자치시인데 원본 122리스트에 아예 없었음"],
    ["전주·창원·청주·김해·아산", "지방 거점도시·학군지인데 누락"],
    ["", ""],
    ["===== 신규 지역 최신 매매 동향 (2026) =====", ""],
    ["대전", "2026.2 보합(0.00%), 서구·유성 상승 vs 원도심 하락"],
    ["광주", "약보합, 광산구(수완·첨단) 상대적 강세"],
    ["세종", "2025 +2.92% 급등 후 2026 상승/하락 반복 불안정"],
    ["전주", "완산구 효자동·덕진구 에코시티 강세"],
    ["창원", "2025 의창구 +11.2% 급등, 기계·방산 산업 호황"],
    ["청주", "오송바이오·오창반도체 배후, 흥덕구 복대동 지웰시티 대장"],
    ["아산", "삼성디스플레이 탕정 배후, 특목고 진학률 전국최상위"],
    ["송도", "대장단지 최고가 대비 33~54% 폭락, GTX-B 예정"],
    ["", ""],
    ["===== 신규 지역 평당가 (2025 기준) =====", ""],
    ["송도국제도시", "3,700만원 (신규 중 최고)"],
    ["세종시", "2,200만원"],
    ["대전 유성구", "1,807만원"],
    ["대전 서구", "1,611만원"],
    ["청주 흥덕구", "1,600만원 (지웰시티 3,350)"],
    ["광주 남구", "1,700만원"],
    ["창원 성산구", "1,500만원"],
    ["전주 완산구", "1,146만원 (신규 중 최저)"],
    ["", ""],
    ["===== 학군점수 실측 교정 =====", ""],
    ["교정 배경", "기존 122리스트 학군점수는 임의배점(0~10)으로 실측 대비 과대"],
    ["실측 공식", "0.35×서울대진학률 + 0.30×특목고진학률 + 0.20×log학원수 + 0.15×배정학군"],
    ["주요 교정", "평촌 9.5→3.63, 수성범어 9.5→3.65, 일산마두 9.0→3.27"],
    ["핵심 발견", "학원가 규모 ≠ 진학성과 (평촌·중계 학원많으나 서울대진학 낮음)"],
    ["", ""],
    ["===== 데이터 출처 =====", ""],
    ["매매지수", "한국부동산원 주간(6월5주)·월간(2026.2~5), 충청투데이·메트로세종"],
    ["평당가", "집품(zippoom)·호갱노노·부킹(buking.kr) 2025~2026"],
    ["학군", "오늘학교 아카데미(academy.prompie.com) 2026"],
    ["", ""],
    ["===== 주의사항 =====", ""],
    ["신뢰도", "신규 지역 대부분 M(범위추정) 등급, 세종·대전서구·유성·송도만 H"],
    ["z-score 기준", "전국 144개 기준 상대평가 — 지방은 절대가격 낮아 하위권"],
    ["추가 정밀화 여지", "창원 나머지 3개구·청주 서원흥덕·전남·경북 미편입"],
]

for r_idx, row in enumerate(method, 1):
    for c_idx, v in enumerate(row, 1):
        c = ws5.cell(row=r_idx, column=c_idx, value=v)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.font = Font(name='Arial', size=10)
        c.border = thin
        if r_idx == 1:
            hstyle(c, color='203864')
        elif '=====' in str(v):
            c.font = Font(name='Arial', size=10, bold=True)
            c.fill = PatternFill('solid', start_color='D9E1F2')

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 72
for r in range(1, len(method)+1):
    ws5.row_dimensions[r].height = 20

out = '/mnt/user-data/outputs/입지등급_확장144_20260703.xlsx'
wb.save(out)
print(f"저장: {out}")
for s in wb.sheetnames:
    print(f"  - {s}")

# CSV도 저장
scored.to_csv('/mnt/user-data/outputs/입지등급_확장144_전체.csv',
              index=False, encoding='utf-8-sig')
print("CSV: 입지등급_확장144_전체.csv")
