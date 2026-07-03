"""
아파트별 입지점수 종합 Excel
- Sheet1: 자치구별 대장 아파트 (25개 구 대표) ★핵심
- Sheet2: 전체 77개 단지 입지점수
- Sheet3: 공식·방법론
"""
import sys
sys.path.insert(0, '/home/claude')
from apartment_scorer import build_apartment_df, compute_apartment_scores, get_district_champions
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

df = compute_apartment_scores(build_apartment_df())
champ = get_district_champions(df)

# 권역 분류
SEOUL_GU = {'종로구','중구','용산구','성동구','광진구','동대문구','중랑구','성북구','강북구',
            '도봉구','노원구','은평구','서대문구','마포구','양천구','강서구','구로구','금천구',
            '영등포구','동작구','관악구','서초구','강남구','송파구','강동구'}
def region_class(gu):
    if gu in SEOUL_GU: return '서울'
    if gu in {'성남시','과천시','수원시','용인시','고양시','광명시','하남시','화성시'}: return '경기'
    if gu in {'송도'}: return '인천'
    if gu.startswith('부산'): return '부산'
    if gu.startswith('대구'): return '대구'
    if gu.startswith('대전'): return '대전'
    return '기타'
champ['권역'] = champ['자치구'].apply(region_class)

wb = Workbook()
thin = Border(*[Side(style='thin', color='D0D0D0')]*4)

def hstyle(c, color='1F4E78'):
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    c.fill = PatternFill('solid', start_color=color)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin

def score_fill(c, s):
    if s < 2.0: c.fill = PatternFill('solid', start_color='9ACD32')
    elif s < 2.5: c.fill = PatternFill('solid', start_color='B5E27C')
    elif s < 3.0: c.fill = PatternFill('solid', start_color='D4F0A0')
    elif s < 3.5: c.fill = PatternFill('solid', start_color='FFF2CC')
    elif s < 4.0: c.fill = PatternFill('solid', start_color='FFCC99')
    else: c.fill = PatternFill('solid', start_color='FF9999')

# ============================================================
# Sheet 1: 자치구별 대장 아파트 (핵심)
# ============================================================
ws1 = wb.active
ws1.title = "전국_지역별_대장아파트"

h1 = ["순위", "권역", "지역", "대장 아파트", "단지\n입지점수", "지역\n점수", "실거래\n평당(만)",
      "세대수", "역세권\n(m)", "재건축성", "초품아", "신뢰도"]
for i, h in enumerate(h1, 1):
    hstyle(ws1.cell(row=1, column=i, value=h))

region_color = {'서울':'FFF2CC','경기':'E2EFDA','인천':'FFE4E1','부산':'F4CCCC',
                '대구':'D9D9D9','대전':'D5E8D4','기타':'EDEDED'}

for idx, r in champ.iterrows():
    vals = [idx+1, r['권역'], r['자치구'], r['단지명'], r['단지입지점수'], round(r['지역점수'],2),
            int(r['실거래평당_만원']), int(r['세대수']), int(r['역세권_m']),
            int(r['재건축성']), '○' if r['초품아']==1 else '-', r['신뢰도']]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=idx+2, column=col, value=v)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col == 2:
            c.fill = PatternFill('solid', start_color=region_color.get(r['권역'], 'FFFFFF'))
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 4:
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 5:
            score_fill(c, r['단지입지점수'])
            c.font = Font(name='Arial', size=9, bold=True)
        if idx == 0:
            c.fill = PatternFill('solid', start_color='FFD700')
            c.font = Font(name='Arial', size=9, bold=True)

widths1 = {'A': 5, 'B': 6, 'C': 10, 'D': 22, 'E': 9, 'F': 8, 'G': 9, 'H': 7, 'I': 7, 'J': 8, 'K': 7, 'L': 7}
for col, w in widths1.items():
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 32
ws1.freeze_panes = 'D2'

# ============================================================
# Sheet 2: 전체 77개 단지
# ============================================================
ws2 = wb.create_sheet("전체102_단지점수")

h2 = ["순위", "단지명", "자치구", "단지\n입지점수", "지역\n점수", "실거래\n평당(만)",
      "세대수", "역세권\n(m)", "준공\n연차", "재건축성", "초품아", "브랜드", "신뢰도", "대장"]
for i, h in enumerate(h2, 1):
    hstyle(ws2.cell(row=1, column=i, value=h), color='2E5C8A')

champ_names = set(champ['단지명'])
for idx, r in df.reset_index(drop=True).iterrows():
    is_champ = r['단지명'] in champ_names
    vals = [idx+1, r['단지명'], r['자치구'], r['단지입지점수'], round(r['지역점수'],2),
            int(r['실거래평당_만원']), int(r['세대수']), int(r['역세권_m']),
            int(r['준공연차']), int(r['재건축성']), '○' if r['초품아']==1 else '-',
            int(r['브랜드']), r['신뢰도'], '🏆' if is_champ else '']
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=idx+2, column=col, value=v)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Arial', size=9)
        c.border = thin
        if col == 2:
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.font = Font(name='Arial', size=9, bold=True) if is_champ else Font(name='Arial', size=9)
        if col == 4:
            score_fill(c, r['단지입지점수'])
            c.font = Font(name='Arial', size=9, bold=True)
        if col == 14 and is_champ:
            c.fill = PatternFill('solid', start_color='FFEB9C')

widths2 = {'A': 5, 'B': 22, 'C': 8, 'D': 9, 'E': 8, 'F': 9, 'G': 7, 'H': 7,
           'I': 6, 'J': 8, 'K': 7, 'L': 7, 'M': 7, 'N': 6}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 32
ws2.freeze_panes = 'C2'
ws2.auto_filter.ref = f"A1:N{len(df)+1}"

# ============================================================
# Sheet 3: 방법론
# ============================================================
ws3 = wb.create_sheet("공식_방법론")
method = [
    ["항목", "내용"],
    ["기준일", "2026년 7월 3일"],
    ["단지 수", "77개 (서울 25개 자치구 전체 커버)"],
    ["", ""],
    ["===== 아파트별 입지점수 공식 =====", ""],
    ["최종 공식", "단지점수 = 지역점수 × 0.5 + 단지자체점수 × 0.5"],
    ["단지자체점수", "5.0 - 4.0 × sigmoid(단지프리미엄)"],
    ["점수 범위", "1.0(극상급) ~ 5.0(하급)"],
    ["", ""],
    ["===== 단지프리미엄 가중치 =====", ""],
    ["역세권 (도보 m)", "30% · 가까울수록 높음"],
    ["실거래 평당가", "20% · 자치구 평균 대비 배수"],
    ["세대수", "15% · 대단지 프리미엄(log)"],
    ["재건축성/신축", "15% · 신축 또는 재건축임박 높음"],
    ["초품아", "10% · 초등학교 도보배정"],
    ["브랜드", "10% · 1군 건설사"],
    ["", ""],
    ["===== 재건축성 배점 기준 =====", ""],
    ["신축(0~5년)", "2점"],
    ["준신축", "3점"],
    ["구축(재건축무관)", "4점"],
    ["재건축 추진초기", "6점"],
    ["재건축 임박(사업시행인가+)", "9점 (은마·압구정현대·잠실주공5 등)"],
    ["", ""],
    ["===== 브랜드 배점 =====", ""],
    ["하이엔드", "10점 (아크로·디에이치·트리마제)"],
    ["1군", "8점 (래미안·자이·푸르지오)"],
    ["2군", "7점 (e편한세상·롯데캐슬)"],
    ["지역/기타", "5점"],
    ["", ""],
    ["===== 서울 자치구별 대장 TOP 5 =====", ""],
    ["1위 서초구", "래미안원베일리 (1.66, 평당 20,117 전국1위)"],
    ["2위 송파구", "헬리오시티 (1.71, 9,510세대 최대)"],
    ["3위 강남구", "래미안대치팰리스 (1.89)"],
    ["4위 마포구", "마포래미안푸르지오 (2.09)"],
    ["5위 용산구", "이촌한가람 (2.17)"],
    ["", ""],
    ["===== 주의사항 =====", ""],
    ["현재 실물 기준", "재건축 미완 단지는 잠재가치 덜 반영(압구정현대 등)"],
    ["예: 압구정현대", "평당 14,714(전국2위)이나 재건축미완+역세권600m로 강남 대장 아님"],
    ["z-score 상대평가", "강서 마곡(평당6,033)이 지역점수 낮아 성북·동대문에 역전"],
    ["확장 방법", "SEED_APARTMENTS에 행 추가 시 자동 재계산"],
    ["", ""],
    ["===== 데이터 출처 =====", ""],
    ["실거래·평당가", "호갱노노·집품(zippoom)·KB부동산·리치고·아파트랭킹 2026"],
    ["재건축 현황", "정비사업 정보몽땅·서울시 고시 2026"],
    ["단지 정보", "세대수·준공·용적률은 각 단지 공개자료"],
]

for r_idx, row in enumerate(method, 1):
    for c_idx, v in enumerate(row, 1):
        c = ws3.cell(row=r_idx, column=c_idx, value=v)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.font = Font(name='Arial', size=10)
        c.border = thin
        if r_idx == 1:
            hstyle(c, color='203864')
        elif '=====' in str(v):
            c.font = Font(name='Arial', size=10, bold=True)
            c.fill = PatternFill('solid', start_color='D9E1F2')

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 62
for r in range(1, len(method)+1):
    ws3.row_dimensions[r].height = 20

out = '/mnt/user-data/outputs/아파트별_입지점수_종합_20260703.xlsx'
wb.save(out)
print(f"저장: {out}")
for s in wb.sheetnames:
    print(f"  - {s}")
